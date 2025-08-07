#!/usr/bin/env python3
import argparse
import logging
import os
import random
import sys
import time
import urllib.parse
from dataclasses import dataclass
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


DEFAULT_MIN_DELAY_SECONDS = 3.5
DEFAULT_MAX_DELAY_SECONDS = 8.0
DEFAULT_NAV_TIMEOUT_SECONDS = 30
DEFAULT_ELEMENT_TIMEOUT_SECONDS = 25
DEFAULT_PROFILE_DIR = os.path.join(os.getcwd(), ".chrome-whatsapp-profile")


@dataclass
class RunConfig:
    excel_path: str
    column_name: Optional[str]
    message_text: str
    image_path: Optional[str]
    video_path: Optional[str]
    min_delay: float
    max_delay: float
    profile_dir: str
    country_code: Optional[str]
    headless: bool


def parse_args() -> RunConfig:
    parser = argparse.ArgumentParser(
        description="Send bulk WhatsApp messages via WhatsApp Web using an Excel contact list."
    )
    parser.add_argument("--excel", required=True, help="Path to Excel file (.xlsx) with contact numbers.")
    parser.add_argument(
        "--column",
        default=None,
        help="Column name in Excel containing phone numbers. If omitted, tries common names else first column.",
    )
    parser.add_argument(
        "--message",
        default=None,
        help="Message text to send. Use literal \\n for newlines or use --message-file for multi-line input.",
    )
    parser.add_argument(
        "--message-file",
        default=None,
        help="Path to a text file whose contents will be used as the message (supports multi-line).",
    )
    parser.add_argument("--image", default=None, help="Optional image file path to attach.")
    parser.add_argument("--video", default=None, help="Optional video file path to attach.")
    parser.add_argument(
        "--min-delay",
        type=float,
        default=DEFAULT_MIN_DELAY_SECONDS,
        help=f"Minimum seconds to wait between messages (default {DEFAULT_MIN_DELAY_SECONDS}).",
    )
    parser.add_argument(
        "--max-delay",
        type=float,
        default=DEFAULT_MAX_DELAY_SECONDS,
        help=f"Maximum seconds to wait between messages (default {DEFAULT_MAX_DELAY_SECONDS}).",
    )
    parser.add_argument(
        "--profile-dir",
        default=DEFAULT_PROFILE_DIR,
        help=f"Chrome user data directory for session persistence (default {DEFAULT_PROFILE_DIR}).",
    )
    parser.add_argument(
        "--country-code",
        default=None,
        help="Digits-only country code to prefix local numbers missing a country code (e.g., 1, 44, 91).",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chrome in headless mode (QR login still requires a profile with existing session).",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging verbosity (default INFO).",
    )

    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    if not os.path.exists(args.excel):
        logging.error("Excel file not found: %s", args.excel)
        sys.exit(2)

    message_text = ""
    if args.message_file:
        if not os.path.exists(args.message_file):
            logging.error("Message file not found: %s", args.message_file)
            sys.exit(2)
        with open(args.message_file, "r", encoding="utf-8") as f:
            message_text = f.read()
    elif args.message is not None:
        message_text = args.message.replace("\\n", "\n")
    else:
        logging.error("You must provide --message or --message-file.")
        sys.exit(2)

    image_path = args.image
    video_path = args.video
    for media in [image_path, video_path]:
        if media and not os.path.exists(media):
            logging.error("Media file not found: %s", media)
            sys.exit(2)

    if args.min_delay < 0 or args.max_delay < 0 or args.max_delay < args.min_delay:
        logging.error("Invalid delay configuration. Ensure 0 <= min <= max.")
        sys.exit(2)

    if args.country_code is not None and not args.country_code.isdigit():
        logging.error("--country-code must be digits only, e.g., 1 or 91.")
        sys.exit(2)

    return RunConfig(
        excel_path=args.excel,
        column_name=args.column,
        message_text=message_text,
        image_path=image_path,
        video_path=video_path,
        min_delay=args.min_delay,
        max_delay=args.max_delay,
        profile_dir=args.profile_dir,
        country_code=args.country_code,
        headless=args.headless,
    )


def _detect_header_and_rows(rows: List[Tuple]) -> Tuple[Optional[List[str]], List[Tuple]]:
    if not rows:
        return None, []
    first = rows[0]
    # If any cell in first row is non-empty string and not purely numeric, treat as header
    header_like = False
    for cell in first:
        if isinstance(cell, str) and cell.strip() != "" and not cell.strip().isdigit():
            header_like = True
            break
    if header_like:
        header = [str(c).strip() if c is not None else "" for c in first]
        data_rows = rows[1:]
        return header, data_rows
    else:
        return None, rows


def load_numbers_from_excel(excel_path: str, column_name: Optional[str]) -> Tuple[List[str], str]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = wb.active
    rows = [tuple(cell for cell in row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        raise ValueError("Excel file is empty.")

    header, data_rows = _detect_header_and_rows(rows)

    selected_idx: int = 0
    selected_name: str = ""

    if header:
        name_to_index = {str(n).strip().lower(): i for i, n in enumerate(header) if n is not None and str(n).strip() != ""}
        if column_name:
            key = column_name.strip().lower()
            if key in name_to_index:
                selected_idx = name_to_index[key]
                selected_name = header[selected_idx]
            else:
                logging.warning("Column '%s' not found. Falling back to autodetect.", column_name)
        if selected_name == "":
            common = [
                "phone",
                "mobile",
                "number",
                "contact",
                "phone_number",
                "mobile_number",
                "whatsapp",
                "whatsapp_number",
            ]
            for c in common:
                if c in name_to_index:
                    selected_idx = name_to_index[c]
                    selected_name = header[selected_idx]
                    break
        if selected_name == "":
            selected_idx = 0
            selected_name = header[0] if header else "Column1"
    else:
        # No header, default to first column
        selected_idx = 0
        selected_name = column_name or "Column1"

    numbers: List[str] = []
    for row in data_rows if header else rows:
        if selected_idx >= len(row):
            continue
        val = row[selected_idx]
        if val is None:
            continue
        s = str(val).strip()
        if not s or s.lower() in {"nan", "none"}:
            continue
        numbers.append(s)

    return numbers, str(selected_name)


def normalize_number(raw: str, country_code: Optional[str]) -> Optional[str]:
    s = raw.strip()
    s = s.replace("(", "").replace(")", "").replace("-", "").replace(" ", "")
    if s.startswith("+"):
        s = s[1:]
    if s.startswith("00"):
        s = s[2:]
    s = "".join(ch for ch in s if ch.isdigit())
    if not s:
        return None

    if country_code:
        if not s.startswith(country_code):
            if s.startswith("0"):
                s = s.lstrip("0")
            if len(s) <= 12:
                s = f"{country_code}{s}"
    return s


def build_driver(profile_dir: str, headless: bool) -> webdriver.Chrome:
    opts = ChromeOptions()
    opts.add_argument(f"--user-data-dir={profile_dir}")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1366,900")
    if headless:
        opts.add_argument("--headless=new")

    try:
        service = ChromeService(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    except WebDriverException as e:
        logging.error("Failed to start Chrome WebDriver: %s", e)
        raise
    return driver


def wait_for_login(driver: webdriver.Chrome) -> None:
    driver.get("https://web.whatsapp.com/")
    wait = WebDriverWait(driver, DEFAULT_NAV_TIMEOUT_SECONDS)

    try:
        wait.until(EC.presence_of_element_located((By.ID, "pane-side")))
        logging.info("WhatsApp Web loaded and logged in.")
        return
    except TimeoutException:
        logging.info("Waiting for QR login... Scan the QR code in the opened browser window.")

    long_wait = WebDriverWait(driver, 300)
    long_wait.until(EC.presence_of_element_located((By.ID, "pane-side")))
    logging.info("Login successful.")


def open_chat_for_number(driver: webdriver.Chrome, e164_digits: str) -> bool:
    url = f"https://web.whatsapp.com/send?phone={urllib.parse.quote(e164_digits)}&app_absent=0"
    driver.get(url)
    wait = WebDriverWait(driver, DEFAULT_NAV_TIMEOUT_SECONDS)

    try:
        wait.until(
            EC.any_of(
                EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']")),
                EC.presence_of_element_located((By.XPATH, "//div[contains(., 'Phone number shared via url is invalid')]")),
                EC.presence_of_element_located((By.XPATH, "//div[contains(., 'invalid phone number')]")),
            )
        )
    except TimeoutException:
        logging.warning("Timeout opening chat for number: %s", e164_digits)
        return False

    try:
        error_elem = driver.find_element(By.XPATH, "//div[contains(., 'invalid') or contains(., 'not on WhatsApp')]")
        if error_elem and error_elem.is_displayed():
            logging.warning("Number not on WhatsApp or invalid: %s", e164_digits)
            return False
    except NoSuchElementException:
        pass

    try:
        driver.find_element(By.XPATH, "//footer//div[@contenteditable='true']")
        return True
    except NoSuchElementException:
        logging.warning("Could not find message box for number: %s", e164_digits)
        return False


def find_visible_editors(driver: webdriver.Chrome) -> List:
    editors = driver.find_elements(By.XPATH, "//div[@contenteditable='true']")
    return [e for e in editors if e.is_displayed()]


def send_multiline_to_active_editor(elem, message: str) -> None:
    lines = message.splitlines()
    for i, line in enumerate(lines):
        elem.send_keys(line)
        if i < len(lines) - 1:
            elem.send_keys(Keys.SHIFT, Keys.ENTER)
    elem.send_keys(Keys.ENTER)


def attach_media_and_send(driver: webdriver.Chrome, message: str, image_path: Optional[str], video_path: Optional[str]) -> bool:
    media_paths = [p for p in [image_path, video_path] if p]
    if not media_paths:
        return False

    wait = WebDriverWait(driver, DEFAULT_ELEMENT_TIMEOUT_SECONDS)

    try:
        clip = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@data-icon='clip']")))
        clip.click()
    except TimeoutException:
        logging.debug("Clip icon not found via data-icon; trying title attribute")
        try:
            clip = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@title='Attach']")))
            clip.click()
        except TimeoutException:
            logging.error("Could not find attachment button.")
            return False

    file_input = None
    try:
        file_input = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='file' and contains(@accept, 'image') and contains(@accept, 'video') ]"))
        )
    except TimeoutException:
        try:
            file_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
            )
        except TimeoutException:
            logging.error("Could not find file chooser input for attachments.")
            return False

    for p in media_paths:
        abs_path = os.path.abspath(p)
        logging.debug("Attaching %s", abs_path)
        file_input.send_keys(abs_path)
        time.sleep(0.8)

    try:
        send_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']/ancestor::div[@role='button']"))
        )
    except TimeoutException:
        logging.error("Send button in media preview not found.")
        return False

    try:
        editors = find_visible_editors(driver)
        if editors:
            send_multiline_to_active_editor(editors[-1], message)
        else:
            logging.debug("No visible editor for caption; will send media without caption and send text after.")
    except Exception as e:
        logging.debug("Failed to type caption: %s", e)

    send_btn.click()

    try:
        if message:
            wait.until(EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']")))
            editors = find_visible_editors(driver)
            if editors:
                send_multiline_to_active_editor(editors[-1], message)
    except Exception as e:
        logging.debug("Failed to send message text after media: %s", e)

    return True


def send_text_only(driver: webdriver.Chrome, message: str) -> bool:
    wait = WebDriverWait(driver, DEFAULT_ELEMENT_TIMEOUT_SECONDS)
    try:
        editor = wait.until(
            EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']"))
        )
    except TimeoutException:
        logging.error("Message editor not found.")
        return False

    send_multiline_to_active_editor(editor, message)
    return True


def human_sleep(min_s: float, max_s: float) -> None:
    delay = random.uniform(min_s, max_s)
    logging.debug("Sleeping for %.2f seconds", delay)
    time.sleep(delay)


def run(config: RunConfig) -> None:
    numbers_raw, used_col = load_numbers_from_excel(config.excel_path, config.column_name)
    logging.info("Loaded %d contacts from column '%s' of %s", len(numbers_raw), used_col, config.excel_path)

    seen = set()
    numbers: List[str] = []
    for raw in numbers_raw:
        norm = normalize_number(raw, config.country_code)
        if not norm:
            continue
        if norm not in seen:
            seen.add(norm)
            numbers.append(norm)

    if not numbers:
        logging.error("No valid phone numbers found after normalization.")
        sys.exit(2)

    driver = build_driver(config.profile_dir, config.headless)
    try:
        wait_for_login(driver)

        total = len(numbers)
        successes = 0
        failures = 0

        for idx, num in enumerate(numbers, start=1):
            logging.info("[%d/%d] Sending to %s", idx, total, num)
            try:
                if not open_chat_for_number(driver, num):
                    failures += 1
                    human_sleep(config.min_delay, config.max_delay)
                    continue

                media_sent = attach_media_and_send(driver, config.message_text, config.image_path, config.video_path)
                if not media_sent:
                    if not send_text_only(driver, config.message_text):
                        failures += 1
                        human_sleep(config.min_delay, config.max_delay)
                        continue

                successes += 1
                human_sleep(config.min_delay, config.max_delay)
            except Exception as e:
                logging.exception("Failed for %s: %s", num, e)
                failures += 1
                human_sleep(config.min_delay, config.max_delay)
                continue

        logging.info("Completed. Success: %d, Failure: %d", successes, failures)
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    cfg = parse_args()
    run(cfg)